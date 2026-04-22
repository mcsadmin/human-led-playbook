"""
parse_spreadsheet.py
====================
Phase 1 of the spreadsheet-to-linear skill.

Reads the task register spreadsheet (.xlsx), resolves all names and labels
against the Linear workspace (using the config from Phase 0), converts
estimated hours to Linear estimate points, and produces a human-reviewable
Markdown document showing exactly what will be created in Linear.

Usage
-----
    python3 parse_spreadsheet.py --spreadsheet task_register.xlsx \
                                  --config ./output/skill_config.json \
                                  --output-dir ./output

The review document (review_document.md) must be approved by the team
before Phase 2 (execution) is run.
"""

from __future__ import annotations
import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from config_schema import (
    SkillConfig, TaskRecord,
    SPREADSHEET_COLUMNS, REQUIRED_COLUMNS,
    normalise_moscow, normalise_status,
    MOSCOW_LABEL_GROUP,
)
from estimate_conversion import convert_hours_to_points, convert_hours_to_label


# ---------------------------------------------------------------------------
# Spreadsheet parsing
# ---------------------------------------------------------------------------

def clean_cell(value) -> str:
    """Convert a cell value to a clean string, handling None and floats."""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def expand_line_breaks(text: str) -> str:
    """Convert HTML <br> tags (used in Markdown tables) to real newlines."""
    return re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)


def parse_spreadsheet(path: str, sheet_name: str = "Task Register") -> list[list[str]]:
    """
    Load the spreadsheet and return all data rows as lists of strings.
    Skips the header row (row 1). Returns rows where column A is non-empty.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Try the named sheet, fall back to first sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
        print(f"  ⚠ Sheet '{sheet_name}' not found. Using '{ws.title}'.")

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # Skip header
        cells = [clean_cell(c) for c in row]
        # Pad to at least 16 columns
        while len(cells) < 16:
            cells.append("")
        task_id = cells[SPREADSHEET_COLUMNS["task_id"]]
        if not task_id:
            continue  # Skip empty rows
        rows.append(cells)

    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Row resolution
# ---------------------------------------------------------------------------

def resolve_row(cells: list[str], config: SkillConfig, scale_name: str) -> TaskRecord:
    """
    Convert a raw spreadsheet row into a fully resolved TaskRecord.
    Populates warnings for any fields that cannot be resolved.
    """
    col = SPREADSHEET_COLUMNS

    task = TaskRecord(
        task_id=cells[col["task_id"]],
        subproject=cells[col["subproject"]],
        task_name=cells[col["task_name"]],
        moscow_raw=cells[col["moscow"]],
        owner_raw=cells[col["owner_r"]],
        accountable_raw=cells[col["accountable_a"]],
        consulted_raw=cells[col["consulted_c"]],
        informed_raw=cells[col["informed_i"]],
        dependencies_raw=cells[col["dependencies"]],
        estimated_hours_raw=cells[col["estimated_hours"]],
        sprint_raw=cells[col["sprint"]],
        end_date_raw=cells[col["end_date"]],
        status_raw=cells[col["status"]],
        context_why=expand_line_breaks(cells[col["context_why"]]),
        task_breakdown=expand_line_breaks(cells[col["task_breakdown"]]),
        acceptance_criteria=expand_line_breaks(cells[col["acceptance_criteria"]]),
    )

    # Validate required fields
    for req in REQUIRED_COLUMNS:
        if not cells[col[req]]:
            task.warnings.append(f"Missing required field: {req}")

    # Resolve MoSCoW
    task.moscow = normalise_moscow(task.moscow_raw)
    if task.moscow_raw and not task.moscow:
        task.warnings.append(f"Unrecognised MoSCoW value: '{task.moscow_raw}'")

    # Resolve owner
    if task.owner_raw:
        member = config.find_member(task.owner_raw)
        if member:
            task.owner_id = member.id
            task.owner_resolved = member.name
        else:
            task.warnings.append(f"Owner '{task.owner_raw}' not found in Linear workspace")

    # Resolve cycle
    if task.sprint_raw:
        cycle = config.find_cycle(task.sprint_raw)
        if cycle:
            task.cycle_id = cycle.id
            task.cycle_resolved = cycle.name
        else:
            if config.cycles:
                task.warnings.append(f"Sprint '{task.sprint_raw}' not found in Linear cycles")
            else:
                task.warnings.append("No Cycles configured in Linear — Sprint assignment skipped")

    # Resolve status
    task.status_name = normalise_status(task.status_raw)
    status_obj = config.find_status(task.status_name)
    if status_obj:
        task.status_id = status_obj.id

    # Convert estimate
    task.estimate_points = convert_hours_to_points(task.estimated_hours_raw, scale_name=scale_name)
    task.estimate_label = convert_hours_to_label(task.estimated_hours_raw, scale_name=scale_name)

    # Parse dependency IDs (comma or semicolon separated)
    if task.dependencies_raw:
        raw_deps = re.split(r"[,;]\s*", task.dependencies_raw)
        task.dependency_ids = [d.strip() for d in raw_deps if d.strip()]

    return task


# ---------------------------------------------------------------------------
# Issue body builder
# ---------------------------------------------------------------------------

def build_issue_body(task: TaskRecord) -> str:
    """
    Render the structured Markdown issue body from a resolved TaskRecord.
    """
    lines = []

    lines += ["### What", task.task_name, ""]

    if task.context_why:
        lines += ["**Context / Why:**", task.context_why, ""]

    if task.task_breakdown:
        lines += ["### Task Breakdown", ""]
        for step in task.task_breakdown.splitlines():
            step = step.strip().lstrip("-").lstrip("*").strip()
            if step:
                lines.append(f"- [ ] {step}")
        lines.append("")

    if task.acceptance_criteria:
        lines += ["### Acceptance Criteria", ""]
        for i, criterion in enumerate(task.acceptance_criteria.splitlines(), 1):
            criterion = criterion.strip().lstrip(f"{i}.").strip()
            if criterion:
                lines.append(f"{i}. {criterion}")
        lines.append("")

    lines += ["---", "*Auto-generated from planning register*", "", "### Planning Metadata"]

    # Accountable — only if different from owner
    if task.accountable_raw and task.accountable_raw.lower() != task.owner_raw.lower():
        lines.append(f"- **Accountable:** {task.accountable_raw}")

    if task.cycle_resolved:
        lines.append(f"- **Sprint / Cycle:** {task.cycle_resolved}")
    elif task.sprint_raw:
        lines.append(f"- **Sprint / Cycle:** {task.sprint_raw} *(unresolved)*")

    if task.estimated_hours_raw:
        pts = f" (= {task.estimate_label} / {task.estimate_points} pts)" if task.estimate_points else ""
        lines.append(f"- **Estimated Hours:** {task.estimated_hours_raw} hrs{pts}")

    if task.dependency_ids:
        lines.append(f"- **Dependencies:** {', '.join(task.dependency_ids)}")

    if task.moscow:
        lines.append(f"- **MoSCoW:** {task.moscow}")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Review document builder
# ---------------------------------------------------------------------------

def build_review_document(
    tasks: list[TaskRecord],
    config: SkillConfig,
    initiative_name: str,
) -> str:
    """
    Generate the human-reviewable Markdown document for Phase 1.
    Groups tasks by sub-project (Linear Project).
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    lines = [
        "# Spreadsheet-to-Linear: Phase 1 Review Document",
        "",
        f"**Generated:** {now}  ",
        f"**Initiative:** {initiative_name or '*(to be created)*'}  ",
        f"**Team:** {config.team.name}",
        "",
        "> **Instructions:** Review the proposed Linear structure below.",
        "> Edit any values directly in this document, then return it for Phase 2 execution.",
        "> Items marked ⚠️ require your attention before proceeding.",
        "",
        "---",
        "",
    ]

    # Summary counts
    total = len(tasks)
    warned = sum(1 for t in tasks if t.has_warnings)
    lines += [
        f"**Summary:** {total} tasks across "
        f"{len(set(t.subproject for t in tasks))} sub-projects. "
        f"{'⚠️ ' + str(warned) + ' tasks have warnings.' if warned else '✅ No warnings.'}",
        "",
        "---",
        "",
    ]

    # Group by subproject
    subprojects: dict[str, list[TaskRecord]] = {}
    for task in tasks:
        subprojects.setdefault(task.subproject, []).append(task)

    for sp_name, sp_tasks in subprojects.items():
        # Derive project dates from task dates
        dates = [t.end_date_raw for t in sp_tasks if t.end_date_raw]
        target_date = max(dates) if dates else ""

        lines += [
            f"## Project: {sp_name}",
            "",
            f"**Linear Object:** Project (Sub-project)  ",
            f"**Target Date:** {target_date or '*(not set)*'}  ",
            f"**Issues:** {len(sp_tasks)}",
            "",
        ]

        for task in sp_tasks:
            warn_flag = " ⚠️" if task.has_warnings else ""
            lines += [
                f"### [{task.task_id}] {task.task_name}{warn_flag}",
                "",
                f"| Field | Value |",
                f"| :--- | :--- |",
                f"| **Issue Title** | `{task.issue_title}` |",
                f"| **Assignee** | {task.owner_resolved or f'⚠️ *{task.owner_raw}* (unresolved)'} |",
                f"| **Status** | {task.status_name} |",
                f"| **MoSCoW Label** | {task.moscow or '*(none)*'} |",
                f"| **Estimate** | {task.estimate_label or '*(none)*'} ({task.estimate_points or '—'} pts) |",
                f"| **Due Date** | {task.end_date_raw or '*(not set)*'} |",
                f"| **Sprint / Cycle** | {task.cycle_resolved or (f'⚠️ *{task.sprint_raw}* (unresolved)' if task.sprint_raw else '*(none)*')} |",
                f"| **Dependencies** | {', '.join(task.dependency_ids) if task.dependency_ids else '*(none)*'} |",
                "",
            ]

            if task.has_warnings:
                lines.append("**⚠️ Warnings:**")
                for w in task.warnings:
                    lines.append(f"- {w}")
                lines.append("")

            lines += [
                "<details>",
                "<summary>Issue body preview (click to expand)</summary>",
                "",
                "```markdown",
                build_issue_body(task),
                "```",
                "",
                "</details>",
                "",
                "---",
                "",
            ]

    lines += [
        "## Approval",
        "",
        "- [ ] I have reviewed all issues above and confirmed the values are correct.",
        "- [ ] All ⚠️ warnings have been resolved.",
        "",
        "**Return this document to trigger Phase 2 (execution).**",
    ]

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Phase 1: Parse spreadsheet and produce Linear review document."
    )
    parser.add_argument("--spreadsheet", required=True, help="Path to the .xlsx task register")
    parser.add_argument("--config", default="./output/skill_config.json",
                        help="Path to skill_config.json from Phase 0")
    parser.add_argument("--sheet", default="Task Register", help="Worksheet name")
    parser.add_argument("--output-dir", default="./output", help="Directory for output files")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load config
    if not Path(args.config).exists():
        print(f"ERROR: Config file not found: {args.config}", file=sys.stderr)
        print("Run preflight.py (Phase 0) first.", file=sys.stderr)
        sys.exit(1)

    print("Phase 1: Parsing spreadsheet...")
    config = SkillConfig.load(args.config)
    print(f"  ✓ Config loaded (team: {config.team.name}, scale: {config.estimate_scale})")

    # Parse spreadsheet
    print(f"  → Reading '{args.spreadsheet}' (sheet: '{args.sheet}')...")
    try:
        rows = parse_spreadsheet(args.spreadsheet, args.sheet)
    except FileNotFoundError:
        print(f"ERROR: Spreadsheet not found: {args.spreadsheet}", file=sys.stderr)
        sys.exit(1)
    print(f"  ✓ {len(rows)} task rows found.")

    # Resolve each row
    tasks = []
    for cells in rows:
        task = resolve_row(cells, config, config.estimate_scale)
        tasks.append(task)

    warned = sum(1 for t in tasks if t.has_warnings)
    if warned:
        print(f"  ⚠ {warned} tasks have unresolved warnings — see review document.")
    else:
        print("  ✓ All tasks resolved without warnings.")

    # Save resolved tasks as JSON (for Phase 2)
    tasks_json_path = output_dir / "resolved_tasks.json"
    with open(tasks_json_path, "w") as f:
        json.dump([
            {
                "task_id": t.task_id,
                "subproject": t.subproject,
                "task_name": t.task_name,
                "issue_title": t.issue_title,
                "issue_body": build_issue_body(t),
                "owner_id": t.owner_id,
                "owner_raw": t.owner_raw,
                "accountable_raw": t.accountable_raw,
                "moscow": t.moscow,
                "estimate_points": t.estimate_points,
                "estimate_label": t.estimate_label,
                "cycle_id": t.cycle_id,
                "status_name": t.status_name,
                "status_id": t.status_id,
                "end_date_raw": t.end_date_raw,
                "dependency_ids": t.dependency_ids,
                "warnings": t.warnings,
            }
            for t in tasks
        ], f, indent=2)
    print(f"  ✓ Resolved tasks saved: {tasks_json_path}")

    # Build and save review document
    review = build_review_document(tasks, config, config.initiative_name)
    review_path = output_dir / "review_document.md"
    review_path.write_text(review, encoding="utf-8")
    print(f"  ✓ Review document saved: {review_path}")

    print("\nPhase 1 complete.")
    print("→ Review the document, resolve any ⚠️ warnings, and return it to trigger Phase 2.")


if __name__ == "__main__":
    main()
